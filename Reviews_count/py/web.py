from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook,Workbook
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")

new_wb = Workbook()
new_ws = new_wb.active

wb = load_workbook(r"D:\Durai\GMB\Reviews_count\Data\Google Ratings & Reviews count urls.xlsx")
ws = wb.active


class GoogleRating:
    def __init__(self):
        pass

    def heading(self,**kwargs):
        new_ws.cell(row=kwargs.get("row_value"), column=2).value = "Rating"
        new_ws.cell(row=kwargs.get("row_value"), column=3).value = "Reviews"

    def file_save(self,**kwargs):
        new_ws.cell(row=kwargs.get('row_value'), column=2).value = driver.find_element(By.CLASS_NAME, "aMPvhf-fI6EEc-KVuj8d").text
        new_ws.cell(row=kwargs.get('row_value'), column=3).value = driver.find_element(By.CLASS_NAME, "Yr7JMd-pane-hSRGPd").text[0:-8]
        new_wb.save(r'D:\Durai\GMB\Reviews_count\save\Google Ratings & Reviews ' + str(kwargs.get('path')) + " " + date + '.xlsx')

    def rating_range(self,**kwargs):
        for r in range(kwargs.get('start'), kwargs.get('end') + 1):
            print(r)
            print(ws.cell(row=r,column=1).value)
            new_ws.cell(row=r, column=1).value = ws.cell(row=r, column=1).value
            if ws.cell(row=r,column=1).value == "STORE NAME":
                GoogleRating.heading(self,row_value=r)
            try:
                if ws.cell(row=r,column=2).value != None:

                    driver.get(ws.cell(row=r, column=2).value)
                    driver.implicitly_wait(10)
                    GoogleRating.file_save(self,row_value=r,path=str(kwargs.get('path')))
            except:
                pass
        driver.close()
        driver.quit()