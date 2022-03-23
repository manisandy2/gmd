from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
'Create a new category'
'Mobile'

"Order online"
""
import time
wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

driver = webdriver.Chrome(r"D:\Durai\Driver\chromedriver.exe")
g = Google(driver=driver)
g.login()

image = r"D:\Durai\GMB\product\Image\17-3-2022-1.jpeg"
url = "https://business.google.com/u/3/products/l/02083316379263937574"
new_category = "Mobile"
Description = "Splash more awesomeness and productivity over your life. Get New A15 Bionic powered Apple iPhone 13 and Upgrade to a premium Apple lifestyle. Visit Poorvika Online or Walk-in to your nearest Poorvika Showroom and get benefits worth upto ₹25,000*! Instant Store Discount of upto ₹8,000*, Cashback upto ₹6000*, upto ₹3000* Exchange Bonus and More! Hurry Now! Valid till 20th March*."
Online_Order = 'https://bit.ly/Apple-iPhone-13'
title = "Apple iPhone 13"

for r in range(400,469):
    print(r)
    time.sleep(10)
    print(ws.cell(row=r,column=2).value)
    driver.get(url=ws.cell(row=r,column=2).value)

    driver.implicitly_wait(10)
    driver.find_element(By.CLASS_NAME,"VfPpkd-vQzf8d").click()
    driver.implicitly_wait(10)
    time.sleep(3)
    driver.find_element(By.CLASS_NAME,"ULhCdf").send_keys(image)
    time.sleep(3)

    driver.find_element(By.ID,'c2').send_keys(title)
    drop_down = driver.find_element(By.CLASS_NAME, "MXWxr")
    print(drop_down.text)
    drop_down.click()
    time.sleep(5)
    try:
        for r2 in driver.find_elements(By.CLASS_NAME, 'VfPpkd-StrnGf-rymPhb-ibnC6b'):
            # print(r2.text)
            if r2.text == new_category:
                # print(r2.text)
                r2.click()
            elif r2.get_attribute("data-value") == "new":
                r2.click()
                time.sleep(2)
                driver.find_element(By.ID, 'c9').send_keys(new_category)
    except:
        driver.find_element(By.ID, 'c9').send_keys(new_category)

    driver.find_element(By.ID,'c24').send_keys(Description)
    for dr in driver.find_elements(By.CLASS_NAME,"WEGjDf"):
        # print(dr.text)
        if dr.text == "Add a button (optional)":
            dr.click()
    time.sleep(3)
    for r1 in driver.find_elements(By.CLASS_NAME,"VfPpkd-StrnGf-rymPhb-b9t22c"):
        if r1.text == "Buy":
            r1.click()
    time.sleep(3)
    driver.find_element(By.ID,"c31").send_keys(Online_Order)
    driver.find_element(By.ID,"c32").click()
    time.sleep(3)
    for r1 in driver.find_elements(By.CLASS_NAME,"VfPpkd-vQzf8d"):
        if r1.text == "Save":
            print("pub")
            r1.click()

driver.quit()
driver.close()
