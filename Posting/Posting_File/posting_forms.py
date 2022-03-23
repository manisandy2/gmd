import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from GMB.Google.Google_login import Google
from selenium import webdriver
from openpyxl import load_workbook

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
wb = load_workbook(r"D:\Durai\GMB\Posting\Data\Posting post urls.xlsx")
ws = wb.active


class GooglePosting:

    def create_post(self):
        for post in driver.find_elements(By.CLASS_NAME, 'U26fgb'):  # This loop is for finding the Add Event po up box
            if post.get_attribute("aria-label") == "Create Post":
                post.click()
                time.sleep(1)
                break

    def move_to_event(self):
        time.sleep(2)
        action = ActionChains(driver)
        action.send_keys(Keys.TAB)
        action.send_keys(Keys.ARROW_RIGHT)
        action.send_keys(Keys.ARROW_RIGHT)
        action.send_keys(Keys.ARROW_RIGHT)
        action.send_keys(Keys.ARROW_RIGHT)
        action.perform()

    def event_click(self):
        for event in driver.find_elements(By.CLASS_NAME, "v5Gsrd"):
            if event.text == "Event":
                event.click()
                break

    def image_upload_form(self):
        time.sleep(2)
        for upload_form in driver.find_elements(By.CLASS_NAME, 'BqN5te'):
            if upload_form.text == 'Add photos or videos':
                upload_form.click()
                time.sleep(1)
                driver.implicitly_wait(30)
            elif upload_form.text == 'Add photos':
                upload_form.click()
                time.sleep(1)
                driver.implicitly_wait(30)

    def image_upload(self,**kwargs):
        test = driver.find_element(By.CLASS_NAME, 'EIlDfe')
        for test1 in test.find_elements(By.CLASS_NAME,'XKSfm-Sx9Kwc'):
            # print('found')
            for test2 in test1.find_elements(By.CLASS_NAME,'XKSfm-Sx9Kwc-bN97Pc'):
                # print('found2')
                for test3 in test2.find_elements(By.CLASS_NAME,'fFW7wc-OEVmcd'):
                    # print('found3')
                    driver.switch_to.frame(test3)
                    # print('switched')
                    time.sleep(1)
                    driver.implicitly_wait(10)
                    for test4 in driver.find_elements(By.ID, 'doclist'):
                        # print('found4')
                        for test5 in test4.find_elements(By.CLASS_NAME,'ge-Fc-cg-Eo-qg'):#<-----------Important--------->
                            # print('found5')
                            driver.implicitly_wait(10)
                            for test6 in test5.find_elements(By.ID,':f'):
                                # print('found6')
                                driver.implicitly_wait(10)
                                for test7 in test6.find_elements(By.XPATH,"//input[@type='file']"):
                                    # print('Close the frame now!')
                                    # test7.send_keys(self.image)
                                    test7.send_keys(kwargs.get('image'))
                                    time.sleep(3)
                                    print("image uploaded")
                                    driver.implicitly_wait(10)

    def post_title(self,**kwargs):
        time.sleep(0.5)
        driver.switch_to.default_content()
        # driver.find_element(By.ID, "c36").send_keys(self.title)
        driver.find_element(By.ID, "c36").send_keys(kwargs.get('title'))
        driver.implicitly_wait(10)

    def add_more_click(self):
        time.sleep(0.5)
        for drop_down in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
            driver.implicitly_wait(30)
            if drop_down.text == 'Add more details (optional)':
                driver.implicitly_wait(30)
                drop_down.click()

    def post_date(self,**kwargs):
        time.sleep(0.5)
        driver.find_element(By.ID, 'c46').clear()
        # driver.find_element(By.ID, 'c46').send_keys(self.end_date)
        driver.find_element(By.ID, 'c46').send_keys(kwargs.get('date'))
        driver.implicitly_wait(30)

    def post_description(self,**kwargs):
        time.sleep(0.5)
        # driver.find_element(By.ID,'c53').send_keys(self.description)
        driver.find_element(By.ID,'c53').send_keys(kwargs.get('description'))
        driver.implicitly_wait(30)

    def none_click(self):
        time.sleep(0.5)
        for drop2 in driver.find_elements(By.CLASS_NAME,'VfPpkd-vQzf8d'):
            if drop2.text == 'None':
                drop2.click()
        driver.implicitly_wait(30)

    def post_call(self,**kwargs):
        time.sleep(0.5)
        for field_call in driver.find_elements(By.CLASS_NAME,'VfPpkd-StrnGf-rymPhb'):
            for field1 in field_call.find_elements(By.CLASS_NAME,"VfPpkd-StrnGf-rymPhb-ibnC6b"):
                driver.implicitly_wait(5)
                if field1.text == kwargs.get('field'):
                    print(kwargs.get('field'))
                    driver.implicitly_wait(30)
                    field1.click()

    def post_published(self):
        time.sleep(0.5)
        for finish1 in driver.find_elements(By.CLASS_NAME, 'bMajjd'):
            for finish in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
                if finish.text == 'Publish':
                    print('Process completed')
                    finish.location_once_scrolled_into_view
                    time.sleep(1.5)
                    driver.implicitly_wait(30)
                    finish1.click()
                    driver.implicitly_wait(30)
                    time.sleep(3)
                    print('Published')


class GooglePostingRun:

    def __init__(self, posting):
        self.posting = posting

    def range_run(self, **kwargs):
        Gg = Google(driver=driver)
        Gg.login()

        print(self.posting.Title)
        driver.implicitly_wait(3)

        for num in range(kwargs.get('start'), kwargs.get('end')):
            print(num)
            print(ws.cell(row=num, column=2).value)
            driver.get(ws.cell(row=num, column=2).value)
            driver.implicitly_wait(3)
            # gp = GooglePosting(self.posting.Img_file, self.posting.Title,
            #                    self.posting.End_date, self.posting.Description,
            #                    self.posting.Field)
            gp = GooglePosting()
            gp.create_post()
            gp.move_to_event()
            gp.event_click()
            gp.image_upload_form()
            gp.image_upload(image=self.posting.Img_file)
            gp.post_title(title=self.posting.Title)
            gp.add_more_click()
            gp.post_date(date=self.posting.End_date)
            gp.post_description(description=self.posting.Description)
            gp.none_click()
            gp.post_call(field=self.posting.Field)
            gp.post_published()
        driver.close()
        driver.quit()





