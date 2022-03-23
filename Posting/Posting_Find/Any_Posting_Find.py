from selenium import webdriver
from openpyxl import load_workbook
import time

from posting.posting_login import google_login

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
wb = load_workbook(r"d:\Durai\posting\excel\posting posts.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

google_login(driver)

time.sleep(7)

driver.implicitly_wait(30)
for num in range(432, 433):
    print(num)
    driver.implicitly_wait(30)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(30)
    data = driver.find_elements_by_class_name("p4HiUc")

    for all_post in data:
        for post in all_post.find_elements_by_class_name("LgQiCc"):
            try:
                for name in post.find_elements_by_class_name("P9ZBeb"):
                    if name.text == "Upto ₹5000 Cashback on OnePlus U1S TVs at Poorvika!":
                        print(name.text)
                        post.find_element_by_class_name("JRtysb").click()

            except:
                pass

driver.quit()
