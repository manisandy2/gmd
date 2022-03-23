from selenium import webdriver
from openpyxl import load_workbook
import time

from posting.posting_login import google_login

driver = webdriver.Chrome(executable_path=r"/Driver/chromedriver.exe")
wb = load_workbook(r"/posting/excel/posting posts.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

google_login(driver)

time.sleep(7)

driver.implicitly_wait(30)
for num in range(100, ws.max_row+1):
    print(num)
    driver.implicitly_wait(30)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(30)
    data = driver.find_elements_by_class_name("p4HiUc")

    for all_post in data:
        for post in all_post.find_elements_by_class_name("LgQiCc"):
            for name in post.find_elements_by_class_name("P9ZBeb"):
                if name.text == "No More Remote Struggles. Make Voice Commands To Your TV!":
                    print(name.text)