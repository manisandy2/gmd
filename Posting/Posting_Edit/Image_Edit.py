from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.by import By
from posting.posting_login import google_login
from selenium.webdriver.common.keys import Keys
from posting.posting_forms import GooglePosting
s = Service(r"D:\Durai\Driver\chromedriver.exe")
driver = webdriver.Chrome(service=s)

wb = load_workbook(r"D:\Durai\posting\excel\posting posts.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

google_login(driver)
time.sleep(3)



# 220 to 256 run
driver.implicitly_wait(10)
for num in range(220, 256):
    print(num)
    driver.implicitly_wait(10)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(10)

    for all_post in driver.find_elements(By.CLASS_NAME,"p4HiUc"):
        for post in all_post.find_elements(By.CLASS_NAME,"LgQiCc"):
            # print(post.text)
            try:
                for name in post.find_elements(By.CLASS_NAME, "P9ZBeb"):
                    if name.text == "Low-Price Alert on realme Watch 2 & Watch 2Pro. Hurry Now!":
                        print(name.text)
                        post.find_element(By.CLASS_NAME, "JRtysb").click()
                        for post_edit in driver.find_elements(By.CLASS_NAME, "z80M1"):
                            if post_edit.text == "Edit":
                                post_edit.click()
                                time.sleep(3)
                                time.sleep(0.5)
                                element = driver.find_element_by_xpath('//*[@id="yDmH0d"]/div[4]/div/div[2]/span/div/div/div/section/div[2]/div/div[1]/div[2]/div[1]/div')
                                driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                                time.sleep(3)
                                for delete_posting in driver.find_elements(By.CLASS_NAME,'VfPpkd-BIzmGd'):
                                    # print(delete_posting.text)
                                    if delete_posting.text == "delete":
                                        delete_posting.click()
                                        time.sleep(3)

                                        for upload_form in driver.find_elements(By.CLASS_NAME, 'BqN5te'):
                                            if upload_form.text == 'Add photos or videos':
                                                upload_form.click()
                                                time.sleep(1)
                                                driver.implicitly_wait(30)
                                            elif upload_form.text == 'Add photos':
                                                upload_form.click()
                                                time.sleep(1)
                                                driver.implicitly_wait(30)

                                        test = driver.find_element(By.CLASS_NAME, 'EIlDfe')
                                        for test1 in test.find_elements(By.CLASS_NAME, 'XKSfm-Sx9Kwc'):
                                            # print('found')
                                            for test2 in test1.find_elements(By.CLASS_NAME, 'XKSfm-Sx9Kwc-bN97Pc'):
                                                # print('found2')
                                                for test3 in test2.find_elements(By.CLASS_NAME, 'fFW7wc-OEVmcd'):
                                                    # print('found3')

                                                    driver.switch_to.frame(test3)
                                                    # print('switched')
                                                    time.sleep(1)
                                                    driver.implicitly_wait(10)
                                                    for test4 in driver.find_elements(By.ID, 'doclist'):
                                                        # print('found4')
                                                        for test5 in test4.find_elements(By.CLASS_NAME,
                                                                                         'ge-Fc-cg-Io-qg'):  # <-----------Important--------->
                                                            # print('found5')
                                                            driver.implicitly_wait(10)
                                                            for test6 in test5.find_elements(By.ID, ':f'):
                                                                # print('found6')
                                                                driver.implicitly_wait(10)
                                                                for test7 in test6.find_elements(By.XPATH,"//input[@type='file']"):
                                                                    # print('Close the frame now!')
                                                                    test7.send_keys(r"D:\Durai\posting\image\18-02-2022-2-1.jpeg")
                                                                    print("image uploaded")
                                                                    time.sleep(7)
                                                                    driver.implicitly_wait(10)

                                        driver.switch_to.default_content()
                                        time.sleep(0.5)

                                        for finish1 in driver.find_elements(By.CLASS_NAME, 'bMajjd'):
                                            for finish in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
                                                if finish.text == 'Publish':
                                                    print('Process completed')
                                                    finish.location_once_scrolled_into_view
                                                    time.sleep(1.5)
                                                    driver.implicitly_wait(30)
                                                    finish1.click()
                                                    driver.implicitly_wait(30)
                                                    time.sleep(3)
                                                    print('Published')

            except:
                pass
