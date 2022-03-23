from selenium import webdriver
from openpyxl import load_workbook
import time
from selenium.webdriver.common.by import By
from posting.posting_login import google_login
from selenium.webdriver.common.keys import Keys
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
wb = load_workbook(r"D:\Durai\posting\excel\posting posts.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

google_login(driver)

time.sleep(7)

driver.implicitly_wait(30)
for num in range(1, ws.max_row+1):
    print(num)
    driver.implicitly_wait(30)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(30)

    for all_post in driver.find_elements(By.CLASS_NAME,"p4HiUc"):
        for post in all_post.find_elements(By.CLASS_NAME,"LgQiCc"):
            try:
                for name in post.find_elements(By.CLASS_NAME,"P9ZBeb"):
                    if name.text == "Get upto ₹11,000* worth Benefits with your New iPhone 13":
                        # print(name.text)
                        post.find_element(By.CLASS_NAME,"JRtysb").click()
                        for post_edit in driver.find_elements(By.CLASS_NAME,"z80M1"):
                            if post_edit.text == "Edit":
                                post_edit.click()
                                time.sleep(3)
                                time.sleep(0.5)
                                driver.find_element(By.ID, 'c12').clear()
                                driver.find_element(By.ID, 'c12').send_keys("""26 Mar 2022""")
                                driver.implicitly_wait(15)

                                time.sleep(3)
                                for finish1 in driver.find_elements(By.CLASS_NAME, 'bMajjd'):
                                    for finish in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
                                        if finish.text == 'Publish':
                                            print('Process completed')
                                            finish.location_once_scrolled_into_view
                                            time.sleep(1.5)
                                            driver.implicitly_wait(30)
                                            finish1.click()
                                            driver.implicitly_wait(30)
                                            time.sleep(4)
                                            print('Published')



            except:
                pass
