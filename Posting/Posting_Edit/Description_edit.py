from selenium import webdriver
from openpyxl import load_workbook
import time
from selenium.webdriver.common.by import By
from posting.posting_login import google_login
from selenium.webdriver.common.keys import Keys
driver = webdriver.Chrome(executable_path=r"/Driver/chromedriver.exe")
wb = load_workbook(r"/posting/excel/posting posts.xlsx")
ws = wb.active
google_login(driver)

time.sleep(7)

driver.implicitly_wait(30)
for num in range(1,200):
    print(num)
    driver.implicitly_wait(30)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(30)
    data = driver.find_elements_by_class_name("p4HiUc")

    for all_post in data:
        for post in all_post.find_elements_by_class_name("LgQiCc"):
            try:
                for name in post.find_elements_by_class_name("P9ZBeb"):
                    if name.text == "Zeb-Smart Cam 101 & Zeb Trimmer Newly Launched @Poorvika":
                        # print(name.text)
                        post.find_element_by_class_name("JRtysb").click()
                        for post_del in driver.find_elements_by_class_name("z80M1"):
                            # print(post_del.text)
                            if post_del.text == "Edit":
                                post_del.click()
                                time.sleep(2)

                                # Description
                                driver.find_element_by_tag_name("textarea").send_keys(Keys.CONTROL, 'a')
                                time.sleep(2)
                                driver.find_element_by_tag_name("textarea").send_keys("""Zebronics recently released two new products that adds more quality to your Home, the Zeb Smart Cam 101 and the Zeb-HT101 Trimmer. The Zeb Smart Cam provides long-lasting Security, while Zeb HT101 Trimmer lets you get styled in seconds! The Smart Cam 101 is available for just ₹1,999*, and the Zeb HT101 Trimmer is available for ₹899*! Order yours now from Poorvika Online or your nearest Poorvika Showroom, Today! T&C*.


                                        Smart Cam : https://poorvika.me/Zebronics_ZEB101
                                        Trimmer : https://poorvika.me/zebronics-trimmer
                                        Chat Now : https://poorvika.me/Whatsapp""")

                                time.sleep(2)
                                for finish1 in driver.find_elements(By.CLASS_NAME, 'bMajjd'):
                                    for finish in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
                                        if finish.text == 'Publish':
                                            print('Process completed')
                                            finish.location_once_scrolled_into_view
                                            time.sleep(1.5)
                                            driver.implicitly_wait(30)
                                            finish1.click()
                                            driver.implicitly_wait(30)
                                            time.sleep(4)
                                            print('Published')

            except:
                pass
