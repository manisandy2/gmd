from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
from selenium.webdriver import ActionChains
from posting_login import google_login
from posting.posting_fields import PostingField1
from datetime import datetime

# # s = Service(r"D:\Durai\Driver\chromedriver 96.0.4\chromedriver.exe")
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
start_time = datetime.now()
print(PostingField1.Title)

wb = load_workbook(r"D:\Durai\posting\excel\posting posts.xlsx")
ws = wb.active
google_login(driver)
time.sleep(7)

for r in range(201,251):
    print(r)
    driver.find_element_by_tag_name("body").send_keys(Keys.CONTROL + 't')
    driver.execute_script("window.open('');")
    time.sleep(3)
    driver.switch_to.window(driver.window_handles[r])
    driver.get(ws.cell(row=r,column=2).value)